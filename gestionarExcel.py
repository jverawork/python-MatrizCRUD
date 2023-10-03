from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.worksheet.table import Table
from collections import Counter
import re

import util_log

#import conexion
camposPaqueteAImprimir = ["TIPO", "NOMBRE_PROCESO", "VARIABLES"]
camposProcesoAImprimir = ["TIPO", "NOMBRE_PROCESO", "VARIABLES", "PARAMETROS"]
camposOperacionesAImprimir = ["SELECT", "INSERT", "DELETE" , "UPDATE", "MERGE"]

tablaDatos = "Tabla1"
hojaBody = "Body"

def formatearCelda(celda):
    celda.font = Font(size=8)
    if celda.column == 10:
        celda.alignment = Alignment(horizontal="center", vertical="center")
    else:
        celda.alignment = Alignment(horizontal="left", vertical="center")

def registrarNombrePaquete(hm_paquete, fila, sheet):
    formatearCelda(sheet.cell(row=fila, column = 1, value = hm_paquete["TIPO"]))
    esquema, _, objeto = hm_paquete["NOMBRE_PROCESO"].partition(".")
    formatearCelda(sheet.cell(row=fila, column = 2, value = esquema))
    formatearCelda(sheet.cell(row=fila, column = 3, value = objeto))

def armarExcel(hm_paquete,archivo, hoja):
    workbook = load_workbook(archivo)
    sheet = workbook[hojaBody]    
    
    table = sheet.tables[tablaDatos]
    table.auto_filter = None

    patron = r"\$?([A-Z]+)\$?(\d+)"
    referencia = table.ref

    filas = re.findall(patron, referencia)
    
    ref_columnaini = filas[0][0]
    ref_filaini = int(filas[0][1])
    ref_columnafin = filas[1][0]
    ref_filafin = int(filas[1][1])
    
    sheet.delete_rows((ref_filaini + 2), ref_filafin-ref_filaini)
    workbook.active
    #print(hm_paquete['VARIABLES'])
    fila = ref_filaini + 1
    listaVariables = [cadena if '.' not in cadena else '.'.join(cadena.split('.')[:2]) for cadena in hm_paquete['VARIABLES']]
    for oper_paquete in set(listaVariables):
        #sheet.insert_rows(fila+1,1)
        registrarNombrePaquete(hm_paquete, fila, sheet)

        try:
            esquema = oper_paquete.split(".")[0]
            objeto = oper_paquete.split(".")[1]
        except Exception as e:
            util_log.logError(hm_paquete["NOMBRE_PROCESO"], hm_paquete, None, None, "gestionarExcel.armarExcel\n"+oper_paquete, e)

        formatearCelda(sheet.cell(row=fila, column = 6, value = esquema))
        formatearCelda(sheet.cell(row=fila, column = 7, value = objeto))
        formatearCelda(sheet.cell(row=fila, column = 9, value = "VARIABLES"))
        contador = sum(1 for elemento in listaVariables if elemento.startswith(esquema+"."+objeto))
        formatearCelda(sheet.cell(row=fila, column = 10, value = contador))      
        sheet.row_dimensions[fila].height = 1125/100
        fila += 1    
    
    for oper_paquete in set(hm_paquete['VARIABLES-DBLINK']):
        #sheet.insert_rows(fila+1,1)
        registrarNombrePaquete(hm_paquete, fila, sheet)
        esquema, _, objeto = oper_paquete.partition(".")
        formatearCelda(sheet.cell(row=fila, column = 6, value = esquema))
        formatearCelda(sheet.cell(row=fila, column = 7, value = objeto))
        formatearCelda(sheet.cell(row=fila, column = 9, value = "VARIABLES"))
        formatearCelda(sheet.cell(row=fila, column = 10, value = hm_paquete['VARIABLES'].count(oper_paquete)))      
        sheet.row_dimensions[fila].height = 1125/100
        fila += 1   
    
    for i in range(len(hm_paquete['LISTA_PROCESOS'])):
        operacion = hm_paquete['LISTA_PROCESOS'][i]
        registrarNombrePaquete(hm_paquete, fila, sheet)
        formatearCelda(sheet.cell(row=fila, column=4, value=operacion["TIPO"]))
        formatearCelda(sheet.cell(row=fila, column=5, value=operacion["NOMBRE_PROCESO"]))
        #PARAMETROS 
        #listaUnicos = set(['.'.join(cadena.split('.')[:-1]) for cadena in operacion['PARAMETROS']])
        listaUnicos = [cadena if '.' not in cadena else '.'.join(cadena.split('.')[:2]) for cadena in operacion['PARAMETROS']]
        esquemaPadre = ''
        #print(operacion["NOMBRE_PROCESO"]," ",operacion["TIPO"]," ",operacion["PARAMETROS"])
        for oper_proceso in set(listaUnicos):
            registrarNombrePaquete(hm_paquete, fila, sheet)
            formatearCelda(sheet.cell(row=fila, column=4, value=operacion["TIPO"]))
            formatearCelda(sheet.cell(row=fila, column=5, value=operacion["NOMBRE_PROCESO"]))
            try:
                esquemaPadre = oper_proceso.split(".")[0]
                objeto = oper_proceso.split(".")[1]
            except Exception as e:
                util_log.logError(operacion["NOMBRE_PROCESO"], operacion, None, listaUnicos, "gestionarExcel.armarExcel parametros\n" + oper_proceso, e)
            formatearCelda(sheet.cell(row=fila, column = 6, value = esquemaPadre))
            formatearCelda(sheet.cell(row=fila, column = 7, value = objeto))            
            formatearCelda(sheet.cell(row=fila, column = 9, value = "PARAMETROS"))
            contador = sum(1 for elemento in listaUnicos if elemento.startswith(esquemaPadre+"."+objeto))          
            formatearCelda(sheet.cell(row=fila, column = 10, value = contador))      
            sheet.row_dimensions[fila].height = 1125/100
            fila += 1   
        #VARIABLES DBLINK
        listaVariables = [[sublista[0].split('.')[0], sublista[1]] for sublista in operacion['VARIABLES-DBLINK']]
        util_log.log(operacion["NOMBRE_PROCESO"], 3, None, None, listaVariables, "gestionarExcel.armarExcel VARIABLES-DBLINK")
        coleccionVariables = Counter(tuple(elemento) for elemento in listaVariables)
        
        for (objeto, dblink), repeticiones in coleccionVariables.items():
                registrarNombrePaquete(hm_paquete, fila, sheet)
                formatearCelda(sheet.cell(row=fila, column=4, value=operacion["TIPO"]))
                formatearCelda(sheet.cell(row=fila, column=5, value=operacion["NOMBRE_PROCESO"]))
                formatearCelda(sheet.cell(row=fila, column = 6, value = "'@"+dblink))
                formatearCelda(sheet.cell(row=fila, column = 7, value = objeto))            
                formatearCelda(sheet.cell(row=fila, column = 9, value = "VARIABLES-DBLINK"))
                formatearCelda(sheet.cell(row=fila, column = 10, value = repeticiones))      
                sheet.row_dimensions[fila].height = 1125/100
                fila += 1  
        #VARIABLES
        listaVariables = [cadena if '.' not in cadena else '.'.join(cadena.split('.')[:2]) for cadena in operacion['VARIABLES']]        
        util_log.log(operacion["NOMBRE_PROCESO"], 4, None, None, listaVariables, "gestionarExcel.armarExcel")    
            
        for oper_proceso in set(listaVariables):
            registrarNombrePaquete(hm_paquete, fila, sheet)
            formatearCelda(sheet.cell(row=fila, column=4, value=operacion["TIPO"]))
            formatearCelda(sheet.cell(row=fila, column=5, value=operacion["NOMBRE_PROCESO"]))
            
            try:
                esquema = oper_proceso.split(".")[0]
                objeto = oper_proceso.split(".")[1]
            except Exception as e:
                util_log.logError(operacion["NOMBRE_PROCESO"], operacion, None, listaVariables, "gestionarExcel.armarExcel\n"+oper_proceso, e)
            formatearCelda(sheet.cell(row=fila, column = 6, value = esquema))
            formatearCelda(sheet.cell(row=fila, column = 7, value = objeto))            
            formatearCelda(sheet.cell(row=fila, column = 9, value = "VARIABLES"))
            contador = sum(1 for elemento in listaVariables if elemento.startswith(esquema+"."+objeto))
            formatearCelda(sheet.cell(row=fila, column = 10, value = contador))      
            sheet.row_dimensions[fila].height = 1125/100
            fila += 1              
        #OPERACIONES INSERT SELECT DELETE EXCUTE UPDATE
        for sentenciaori, tablas in operacion["OPERACIONES"].items():        
            for tablaori in set(tablas):                 
                sentencia = sentenciaori
                tabla = tablaori
                esquema, objeto = "", ""
                registrarNombrePaquete(hm_paquete, fila, sheet)
                formatearCelda(sheet.cell(row=fila, column=4, value=operacion["TIPO"]))
                formatearCelda(sheet.cell(row=fila, column=5, value=operacion["NOMBRE_PROCESO"]))
                noHayPunto = False
                if '.NEXTVAL' in tabla:
                    tabla = tabla.replace('.NEXTVAL','')                
                if '@' in tabla:
                    objeto, _, esquema = tabla.partition("@")
                    esquema = "'@"+esquema
                    sentencia = sentencia + "-DBLINK"
                else:
                    if "." in tabla:
                        esquema, _, objeto = tabla.partition(".")
                    else:
                        objeto = tabla
                        noHayPunto = True
                #if sentencia == "SELECT":
                    #print(tabla)
                if sentencia != "EXECUTE":
                    formatearCelda(sheet.cell(row=fila, column = 6, value = esquema))
                    formatearCelda(sheet.cell(row=fila, column = 7, value = objeto)) 
                else:
                    paquete, _, subproceso = objeto.partition(".")
                    if noHayPunto:
                        paquete , subproceso = "", objeto
                    formatearCelda(sheet.cell(row=fila, column = 7, value = paquete)) 
                    formatearCelda(sheet.cell(row=fila, column = 8, value = subproceso))                      
                formatearCelda(sheet.cell(row=fila, column = 9, value = sentencia))
                formatearCelda(sheet.cell(row=fila, column = 10, value = tablas.count(tablaori))) 
              
                           
                sheet.row_dimensions[fila].height = 1125/100
                fila += 1

    #sheet.delete_rows((ref_filaini + 1), 1)
    table.ref = f"${ref_columnaini}${ref_filaini}:${ref_columnafin}${fila-1}"
    #print(table.ref)
    table.auto_filter = None
    #print(table.ref)
   
    #python_private_range = DefinedName('tablaDatos', attr_text=f'{sheet}!{cell2}', localSheetId=sheetid)
    #workbook.defined_names.append(python_private_range)
    workbook.save(archivo)
    workbook.close()

#
    #for sentencia, tablas in hm_paquete["OPERACIONES"].items():        
    #    for tabla in tablas: 
    #        sheet.cell(row=row_num, column=1, value=hm_paquete["NOMBRE_PROCESO"])            
    #        sheet.cell(row=row_num, column=3, value=hm_paquete["TIPO"])
    #        sheet.cell(row=row_num, column=4, value=tabla)        
    #        sheet.cell(row=row_num, column=5, value=sentencia)
    #        row_num += 1
    #tablaExcel = workbook.defined_names["tblTablaBDD"]
    #tablaExcel.attr_text = f"bdd!$A$1:$E${row_num-1}"
#
    #fila_inicio = 1
    #for fila, (clave, valores) in enumerate(datos.items(), start=fila_inicio):
    #    hoja.cell(row=fila, column=1, value=clave)
    #    for col, valor in enumerate(valores, start=2):
    #        hoja.cell(row=fila, column=col, value=valor)
    #
    #workbook.save(archivo) 

    #datos = conexion.consultar_matrizobjcomparti2iessbiess_t()
    #for dato in datos:
    #    print(dato)

def armarExcelconDataFrame(hm_paquete,archivo, hoja):
    workbook = load_workbook(archivo)
    
    #row_num = 2    
    #col_num = 5 
    #sheet = workbook["Hoja1"]
    #alineacion = Alignment(text_rotation="45")
    #side_obj = Side(border_style='hair')
    #bordes = Border(left=side_obj, right=side_obj, top=side_obj, bottom=side_obj)
    #sheet.delete_rows(row_num+1, amount=sheet.max_row)
    #for table, ops in tablas.items():
    #    unique_ops = list(set(ops))
    #    operations_str = "\n".join(unique_ops)
    #    sheet.cell(row=row_num, column=col_num, value=table).alignment = alineacion
    #    sheet.cell(row=row_num, column=col_num, value=table).border = bordes
    #    sheet.cell(row=row_num+1, column=col_num, value=operations_str)
    #    col_num += 1
    #workbook.save("operations.xlsx") 

    #sheet = workbook[hoja]
    #row_num = 2
    #sheet.delete_rows(row_num+1, amount=sheet.max_row)

    dfPaquete = pd.DataFrame({clave: hm_paquete[clave] for clave in camposPaqueteAImprimir})
    dfPaquete["TIPOVAR"] = "VARIABLES"
    print(dfPaquete)
    #for proceso in hm_paquete["LISTA_PROCESOS"]:
    #    dfProceso = pd.DataFrame({clave: proceso[clave] for clave in camposProcesoAImprimir})
        #print(dfProceso)
    #pd.DataFrame.from_dict(hm_paquete)
    #df.to_excel(archivo, index=False)

    #for i in range(len(hm_paquete['LISTA_PROCESOS'])):
    #    operacion = hm_paquete['LISTA_PROCESOS'][i]
    #    #print(operacion["NOMBRE_PROCESO"])
    #    #print(operacion["OPERACIONES"])
    #    for sentencia, tablas in operacion["OPERACIONES"].items():        
    #        for tabla in tablas: 
    #            sheet.insert_rows(row_num, amount=1)
    #            sheet.cell(row=row_num, column=1, value=hm_paquete["NOMBRE_PROCESO"])
    #            sheet.cell(row=row_num, column=2, value=operacion["NOMBRE_PROCESO"])
    #            sheet.cell(row=row_num, column=3, value=operacion["TIPO"])
    #            sheet.cell(row=row_num, column=4, value=tabla)        
    #            sheet.cell(row=row_num, column=5, value=sentencia)
    #            row_num += 1
#
    #for sentencia, tablas in hm_paquete["OPERACIONES"].items():        
    #    for tabla in tablas: 
    #        sheet.cell(row=row_num, column=1, value=hm_paquete["NOMBRE_PROCESO"])            
    #        sheet.cell(row=row_num, column=3, value=hm_paquete["TIPO"])
    #        sheet.cell(row=row_num, column=4, value=tabla)        
    #        sheet.cell(row=row_num, column=5, value=sentencia)
    #        row_num += 1
    #tablaExcel = workbook.defined_names["tblTablaBDD"]
    #tablaExcel.attr_text = f"bdd!$A$1:$E${row_num-1}"
#
    #fila_inicio = 1
    #for fila, (clave, valores) in enumerate(datos.items(), start=fila_inicio):
    #    hoja.cell(row=fila, column=1, value=clave)
    #    for col, valor in enumerate(valores, start=2):
    #        hoja.cell(row=fila, column=col, value=valor)
    #
    #workbook.save(archivo) 

    #datos = conexion.consultar_matrizobjcomparti2iessbiess_t()
    #for dato in datos:
    #    print(dato)

def estiloTabla():
    # Definir un estilo para la cabecera de la tabla
    style = TableStyleInfo(
        name="TableStyleMedium2", 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False)
    return style

    #cabecera_style = NamedStyle(name="Cabecera")
    #cabecera_style.font = Font(bold=True)
    #cabecera_style.border = Border(bottom=Side(style="thin"))
    #cabecera_style.alignment = Alignment(horizontal="center")
    #return cabecera_style